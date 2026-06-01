"""
Phase 2 + Phase 3 数据合并 — 共享工具函数
"""

import re
import unicodedata
from pathlib import Path

import openpyxl


# ── 字段映射 ──────────────────────────────────────────────
# Excel 列名 → CSV 列名
FIELD_MAP = {
    "mediac": "期刊",
    "years": "年份",
    "vol": "卷",
    "num": "期",
    "titlec": "题目",
    "showwriter": "作者",
    "showorgan": "作者机构",
    "pagecount": "页数",
    "keywordc": "主题词",
}

CSV_COLUMNS = list(FIELD_MAP.values())


# ── 文本规范化 ────────────────────────────────────────────
def normalize_text(s: str) -> str:
    """
    规范化文本用于比较：
    - 全角 → 半角
    - 智能引号 → ASCII 引号
    - 空白规范化
    """
    if not s:
        return ""
    # NFKC 规范化（全角→半角等）
    s = unicodedata.normalize("NFKC", s)
    # 智能引号替换
    s = s.replace("“", '"').replace("”", '"')
    s = s.replace("‘", "'").replace("’", "'")
    # 中文标点统一
    s = s.replace("：", ":").replace("，", ",")
    s = s.replace("；", ";").replace("！", "!")
    s = s.replace("？", "?")
    # 空白规范化
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ── 文件名安全化 ──────────────────────────────────────────
_UNSAFE_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_MULTI_UNDERSCORE = re.compile(r"_{2,}")


def sanitize_filename(name: str, max_bytes: int = 200) -> str:
    """
    将字符串转化为安全的文件名：
    - 移除/替换不安全字符
    - 中文标点 → _
    - 连续下划线压缩
    - UTF-8 字节长度截断
    """
    # 替换不安全 ASCII 字符和中文标点
    name = _UNSAFE_CHARS.sub("_", name)
    name = name.replace("：", "_").replace("“", "_").replace("”", "_")
    name = name.replace("‘", "_").replace("’", "_")
    name = name.replace("《", "_").replace("》", "_")
    name = name.replace("（", "(").replace("）", ")")
    # 连续下划线压缩
    name = _MULTI_UNDERSCORE.sub("_", name)
    # 首尾下划线和空白
    name = name.strip("_ ")
    # UTF-8 截断
    encoded = name.encode("utf-8")
    if len(encoded) > max_bytes:
        # 在字符边界截断
        truncated = encoded[:max_bytes].decode("utf-8", errors="ignore")
        name = truncated.rstrip("_ ")
    return name


# ── 标题模糊匹配 ──────────────────────────────────────────
def _lcs_ratio(a: str, b: str) -> float:
    """计算两个字符串的最长公共子串比率（0-1）"""
    if not a or not b:
        return 0.0
    # 使用较短字符串长度作为分母
    min_len = min(len(a), len(b))
    max_len = max(len(a), len(b))
    if min_len == 0:
        return 0.0

    # 动态规划求 LCS
    prev = [0] * (len(b) + 1)
    max_lcs = 0
    for i in range(1, len(a) + 1):
        curr = [0] * (len(b) + 1)
        for j in range(1, len(b) + 1):
            if a[i - 1] == b[j - 1]:
                curr[j] = prev[j - 1] + 1
                max_lcs = max(max_lcs, curr[j])
        prev = curr

    return max_lcs / min_len


def fuzzy_match_title(pdf_stem: str, titlec: str, threshold: float = 0.85) -> float:
    """
    比较 PDF 文件名（无后缀）与 Excel titlec 的匹配度。
    返回 0-1 的匹配分数。
    """
    a = normalize_text(pdf_stem)
    b = normalize_text(titlec)

    # 精确匹配
    if a == b:
        return 1.0

    # 移除所有非文字字符后比较
    a_clean = re.sub(r"[^a-zA-Z0-9一-鿿]", "", a)
    b_clean = re.sub(r"[^a-zA-Z0-9一-鿿]", "", b)
    if a_clean == b_clean:
        return 0.99

    # LCS 比率
    ratio = _lcs_ratio(a_clean, b_clean)
    return ratio if ratio >= threshold else 0.0


# ── Excel 读取 ────────────────────────────────────────────
def load_excel_metadata(path: str | Path) -> list[dict]:
    """
    读取 Excel 文件并提取目标字段。
    返回字典列表，每个字典的键为 CSV 列名。
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    # 获取列名到索引的映射
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_indices = {}
    for excel_col, csv_col in FIELD_MAP.items():
        if excel_col in headers:
            col_indices[csv_col] = headers.index(excel_col)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for csv_col, idx in col_indices.items():
            val = row[idx] if idx < len(row) else None
            record[csv_col] = val if val is not None else ""
        # 只保留至少有题目的行
        if record.get("题目"):
            rows.append(record)

    wb.close()
    return rows


def load_temp_xlsx(path: str | Path) -> list[dict]:
    """
    读取 temp.xlsx，返回字典列表。
    temp.xlsx 缺少 showwriter/showorgan/pagecount/keywordc，
    这些字段设为空字符串，后续由爬取填充。
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # temp.xlsx 可用字段映射
    temp_field_map = {
        "mediac": "期刊",
        "years": "年份",
        "vol": "卷",
        "num": "期",
        "titlec": "题目",
    }

    col_indices = {}
    for excel_col, csv_col in temp_field_map.items():
        if excel_col in headers:
            col_indices[csv_col] = headers.index(excel_col)

    # 额外提取 lngid 和 链接（用于爬取）
    extra_cols = {"lngid": None, "链接": None}
    for col_name in extra_cols:
        if col_name in headers:
            extra_cols[col_name] = headers.index(col_name)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for csv_col, idx in col_indices.items():
            val = row[idx] if idx < len(row) else None
            record[csv_col] = val if val is not None else ""

        # 提取 lngid 和 URL
        lngid_idx = extra_cols.get("lngid")
        url_idx = extra_cols.get("链接")
        record["_lngid"] = row[lngid_idx] if lngid_idx is not None else ""
        record["_url"] = row[url_idx] if url_idx is not None else ""

        # 缺失字段设为空
        record.setdefault("作者", "")
        record.setdefault("作者机构", "")
        record.setdefault("页数", "")
        record.setdefault("主题词", "")

        # 只保留至少有题目的行，且排除目录页
        title = record.get("题目", "")
        lngid = record.get("_lngid", "")
        if title and lngid != "FXYJ2025006013":
            rows.append(record)

    wb.close()
    return rows
