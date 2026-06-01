#!/usr/bin/env python3
"""从 Excel 文件读取期刊信息并更新报告"""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def load_journal_mapping():
    """从三个 Excel 文件加载期刊信息"""
    journal_mapping = {}

    excel_files = [
        ("raw/4_中国法学_论文信息.xlsx", "中国法学"),
        ("raw/8_中国社会科学_论文信息.xlsx", "中国社会科学"),
        ("raw/11_法学研究_论文信息.xlsx", "法学研究")
    ]

    for excel_file, journal_name in excel_files:
        try:
            print(f"读取 {excel_file}...")
            wb = load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active

            # 读取表头
            headers = [cell.value for cell in ws[1]]

            # 找到标题列（篇名）
            title_col_idx = None
            for idx, col in enumerate(headers):
                if col in ['篇名', '标题', '题名', 'title', '论文标题']:
                    title_col_idx = idx
                    print(f"  找到标题列: {col} (索引 {idx})")
                    break

            if title_col_idx is not None:
                count = 0
                # 只读取标题列，提高速度
                for row in ws.iter_rows(min_row=2, max_col=title_col_idx+1, values_only=True):
                    if len(row) > title_col_idx and row[title_col_idx]:
                        title = str(row[title_col_idx]).strip()
                        if title and title != 'None':
                            journal_mapping[title] = journal_name
                            count += 1
                print(f"  从 {journal_name} 提取了 {count} 个标题")
            else:
                print(f"  警告: 未找到标题列，可用列: {headers[:10]}")

            wb.close()

        except Exception as e:
            print(f"读取 {excel_file} 失败: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n总计加载 {len(journal_mapping)} 个标题-期刊映射")
    return journal_mapping

def match_journal_from_dirs(paper_filename):
    """从法学三大刊论文目录匹配期刊"""
    import os

    # 移除 .pdf 后缀
    title = paper_filename.replace('.pdf', '').strip()

    # 检查三个期刊目录
    journal_dirs = [
        ("法学三大刊论文/中国法学", "中国法学"),
        ("法学三大刊论文/中国社会科学", "中国社会科学"),
        ("法学三大刊论文/法学研究", "法学研究")
    ]

    for dir_path, journal_name in journal_dirs:
        if os.path.exists(dir_path):
            # 列出目录中的所有 PDF 文件
            for filename in os.listdir(dir_path):
                if not filename.endswith('.pdf'):
                    continue

                file_title = filename.replace('.pdf', '').strip()

                # 直接匹配
                if title == file_title:
                    return journal_name

                # 模糊匹配（移除特殊字符）
                clean_title = title.replace('_', '').replace('-', '').replace(' ', '').replace('：', '').replace(':', '')
                clean_file = file_title.replace('_', '').replace('-', '').replace(' ', '').replace('：', '').replace(':', '')
                if clean_title == clean_file:
                    return journal_name

                # 包含匹配（标题在文件名中，或文件名在标题中）
                if title in file_title or file_title in title:
                    return journal_name

    return None

def match_journal(paper_filename, journal_mapping):
    """根据文件名匹配期刊"""
    # 先尝试从目录匹配
    journal = match_journal_from_dirs(paper_filename)
    if journal:
        return journal

    # 移除 .pdf 后缀
    title = paper_filename.replace('.pdf', '').strip()

    # 直接匹配
    if title in journal_mapping:
        return journal_mapping[title]

    # 尝试模糊匹配（移除特殊字符）
    clean_title = title.replace('_', '').replace('-', '').replace(' ', '')
    for mapped_title, journal in journal_mapping.items():
        clean_mapped = mapped_title.replace('_', '').replace('-', '').replace(' ', '')
        if clean_title == clean_mapped:
            return journal

    # 尝试包含匹配
    for mapped_title, journal in journal_mapping.items():
        if mapped_title in title or title in mapped_title:
            return journal

    return "未知期刊"

def main():
    results_file = Path(__file__).parent.parent / "results" / "cross-review-enhanced-analysis.json"
    output_file = Path(__file__).parent.parent / "results" / "round2-full-report.md"

    # 加载期刊映射
    print("正在加载期刊信息...")
    journal_mapping = load_journal_mapping()

    with open(results_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 提取所有论文的第二轮数据
    papers = []
    matched_count = 0
    unmatched_papers = []

    for paper_data in data["papers"]:
        paper_name = paper_data["paper"]
        dimensions = paper_data["dimensions"]

        # 计算第二轮六维平均分
        round2_scores = []
        round2_stds = []
        for dim_name, dim_data in dimensions.items():
            if "round2_mean" in dim_data:
                round2_scores.append(dim_data["round2_mean"])
                round2_stds.append(dim_data["round2_std"])

        if round2_scores:
            # 匹配期刊
            journal = match_journal(paper_name, journal_mapping)
            if journal != "未知期刊":
                matched_count += 1
            else:
                unmatched_papers.append(paper_name)

            # 提取论文标题（保留原始文件名作为标题）
            title = paper_name.replace('.pdf', '')

            papers.append({
                "title": title,
                "journal": journal,
                "round2_mean": sum(round2_scores) / len(round2_scores),
                "round2_avg_std": sum(round2_stds) / len(round2_stds),
                "filename": paper_name
            })

    print(f"\n匹配结果: {matched_count}/{len(papers)} 篇论文匹配到期刊")
    if unmatched_papers:
        print(f"\n未匹配的论文 ({len(unmatched_papers)} 篇):")
        for p in unmatched_papers[:10]:
            print(f"  - {p}")
        if len(unmatched_papers) > 10:
            print(f"  ... 还有 {len(unmatched_papers) - 10} 篇")

    # 按平均分降序排序
    papers.sort(key=lambda x: x["round2_mean"], reverse=True)

    # 生成 Markdown 报告
    with open(output_file, "w", encoding="utf-8") as f:
        f.write("# 第二轮交叉评审完整报告\n\n")
        f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"**总论文数**: {len(papers)} 篇\n\n")

        # 统计信息
        scores = [p["round2_mean"] for p in papers]
        stds = [p["round2_avg_std"] for p in papers]

        f.write("## 整体统计\n\n")
        f.write(f"- **平均分**: {sum(scores) / len(scores):.2f}\n")
        f.write(f"- **中位数**: {sorted(scores)[len(scores) // 2]:.2f}\n")
        f.write(f"- **最高分**: {max(scores):.2f}\n")
        f.write(f"- **最低分**: {min(scores):.2f}\n")
        f.write(f"- **平均标准差**: {sum(stds) / len(stds):.2f}\n")
        f.write(f"- **期刊匹配率**: {matched_count}/{len(papers)} ({matched_count/len(papers)*100:.1f}%)\n\n")

        # 分数区间统计
        f.write("## 分数区间分布\n\n")
        bins = [
            (90, 100, "90-100"),
            (85, 90, "85-90"),
            (80, 85, "80-85"),
            (75, 80, "75-80"),
            (70, 75, "70-75"),
            (60, 70, "60-70"),
            (0, 60, "<60")
        ]

        f.write("| 分数区间 | 论文数 | 占比 |\n")
        f.write("|---------|--------|------|\n")
        for low, high, label in bins:
            count = sum(1 for s in scores if low <= s < high)
            percentage = count / len(scores) * 100
            f.write(f"| {label} | {count} 篇 | {percentage:.1f}% |\n")

        f.write("\n")

        # 标准差区间统计
        f.write("## 标准差区间分布\n\n")
        std_bins = [
            (0, 3, "<3 (高一致性)"),
            (3, 5, "3-5 (中等一致性)"),
            (5, 8, "5-8 (低一致性)"),
            (8, 100, ">8 (分歧显著)")
        ]

        f.write("| 标准差区间 | 论文数 | 占比 |\n")
        f.write("|-----------|--------|------|\n")
        for low, high, label in std_bins:
            count = sum(1 for s in stds if low <= s < high)
            percentage = count / len(stds) * 100
            f.write(f"| {label} | {count} 篇 | {percentage:.1f}% |\n")

        f.write("\n")

        # 完整论文列表
        f.write("## 完整论文列表\n\n")
        f.write("| 排名 | 平均分 | 标准差 | 期刊 | 论文标题 |\n")
        f.write("|------|--------|--------|------|----------|\n")

        for i, paper in enumerate(papers, 1):
            f.write(f"| {i} | {paper['round2_mean']:.2f} | {paper['round2_avg_std']:.2f} | {paper['journal']} | {paper['title']} |\n")

        f.write("\n")

        # 分组统计
        f.write("## 分组统计\n\n")

        # 按分数分组
        f.write("### 按分数分组\n\n")
        f.write("#### 优秀论文（≥80 分）\n\n")
        excellent = [p for p in papers if p["round2_mean"] >= 80]
        f.write(f"共 {len(excellent)} 篇\n\n")
        f.write("| 排名 | 平均分 | 标准差 | 期刊 | 论文标题 |\n")
        f.write("|------|--------|--------|------|----------|\n")
        for i, p in enumerate(excellent, 1):
            f.write(f"| {i} | {p['round2_mean']:.2f} | {p['round2_avg_std']:.2f} | {p['journal']} | {p['title']} |\n")
        f.write("\n")

        f.write("#### 良好论文（70-80 分）\n\n")
        good = [p for p in papers if 70 <= p["round2_mean"] < 80]
        f.write(f"共 {len(good)} 篇\n\n")

        f.write("#### 及格论文（60-70 分）\n\n")
        pass_papers = [p for p in papers if 60 <= p["round2_mean"] < 70]
        f.write(f"共 {len(pass_papers)} 篇\n\n")

        f.write("#### 不及格论文（<60 分）\n\n")
        fail = [p for p in papers if p["round2_mean"] < 60]
        f.write(f"共 {len(fail)} 篇\n\n")
        f.write("| 排名 | 平均分 | 标准差 | 期刊 | 论文标题 |\n")
        f.write("|------|--------|--------|------|----------|\n")
        for p in fail:
            rank = papers.index(p) + 1
            f.write(f"| {rank} | {p['round2_mean']:.2f} | {p['round2_avg_std']:.2f} | {p['journal']} | {p['title']} |\n")
        f.write("\n")

        # 按标准差分组
        f.write("### 按标准差分组\n\n")

        f.write("#### 高一致性（std < 3）\n\n")
        high_consistency = [p for p in papers if p["round2_avg_std"] < 3]
        f.write(f"共 {len(high_consistency)} 篇\n\n")
        f.write("| 排名 | 平均分 | 标准差 | 期刊 | 论文标题 |\n")
        f.write("|------|--------|--------|------|----------|\n")
        for p in high_consistency:
            rank = papers.index(p) + 1
            f.write(f"| {rank} | {p['round2_mean']:.2f} | {p['round2_avg_std']:.2f} | {p['journal']} | {p['title']} |\n")
        f.write("\n")

        f.write("#### 分歧显著（std > 8）\n\n")
        high_divergence = [p for p in papers if p["round2_avg_std"] > 8]
        f.write(f"共 {len(high_divergence)} 篇\n\n")
        f.write("| 排名 | 平均分 | 标准差 | 期刊 | 论文标题 |\n")
        f.write("|------|--------|--------|------|----------|\n")
        for p in high_divergence:
            rank = papers.index(p) + 1
            f.write(f"| {rank} | {p['round2_mean']:.2f} | {p['round2_avg_std']:.2f} | {p['journal']} | {p['title']} |\n")
        f.write("\n")

        # 期刊统计
        f.write("## 期刊统计\n\n")
        journal_stats = {}
        for p in papers:
            journal = p["journal"]
            if journal not in journal_stats:
                journal_stats[journal] = {
                    "count": 0,
                    "scores": [],
                    "stds": []
                }
            journal_stats[journal]["count"] += 1
            journal_stats[journal]["scores"].append(p["round2_mean"])
            journal_stats[journal]["stds"].append(p["round2_avg_std"])

        # 按论文数量降序排序
        sorted_journals = sorted(journal_stats.items(), key=lambda x: x[1]["count"], reverse=True)

        f.write("| 期刊 | 论文数 | 平均分 | 平均标准差 |\n")
        f.write("|------|--------|--------|------------|\n")
        for journal, stats in sorted_journals:
            avg_score = sum(stats["scores"]) / len(stats["scores"])
            avg_std = sum(stats["stds"]) / len(stats["stds"])
            f.write(f"| {journal} | {stats['count']} | {avg_score:.2f} | {avg_std:.2f} |\n")

        f.write("\n---\n\n")
        f.write("*本报告由 SocialEval 系统自动生成*\n")

    print(f"\n报告已生成: {output_file}")
    print(f"总计 {len(papers)} 篇论文")

if __name__ == "__main__":
    main()
